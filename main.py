import datetime
import os
import time
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import smtplib

from openpyxl.styles.numbers import BUILTIN_FORMATS


driver = webdriver.Firefox()
driver.get("https://yandex.ru/")
driver.find_element_by_link_text("USD MOEX").click()
driver.back()
driver.find_element_by_link_text("EUR MOEX").click()
driver.back()

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
             'Chrome/103.0.0.0 Safari/537.36'


FN_RESULT = '{}_result.xlsx'

USD_URL = 'https://yandex.ru/news/quotes/2002.html'
EUR_URL = 'https://yandex.ru/news/quotes/2000.html'

PAUSE_NEXT = 10  # пауза перед следующей валютой (чтобы снизить вероятность появления капчи)
RATIO_ROUND_DIGIT = 3  # сколько символов в соотношеии цены


def get_response(url, headers=None, max_tries=3, pause_next=10):
    for try_num in range(1, max_tries + 1):
        try:
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                return response.text
        except Exception as e:
            if try_num >= 2:
                print(f'Ошибка получения ответа при запросе "{url}": {e}')

        time.sleep(pause_next)

    print('Не удалось получить ответ!')
    return None


def pars_currency(response):
    soup = BeautifulSoup(response, 'lxml')

    currencies = []
    rows = soup.find_all("div", {"class": "news-stock-table__row"})
    if rows:
        for row in rows[1:]:
            try:
                classes = ",".join(row.get('class'))
                if 'negative' in classes:
                    sign = -1
                else:
                    sign = 1
                date = row.find_all("div", {"class": "news-stock-table__cell"})[0].text
                value = float(row.find_all("div",
                                           {"class": "news-stock-table__cell"})[1].text.replace(',', '.'))
                change = sign * float(row.find_all("div",
                                                   {"class": "news-stock-table__cell"})[2].text.replace(',', '.'))
            except:
                continue
            if date and value and change:
                currencies.append({"date": date, "value": value, "change": change})

    if currencies:
        print(f'Количество собранных значений = {len(currencies)}')
    else:
        if 'SmartCaptcha' in response:
            print('Обнаружен запрос капчи!')
        else:
            print('Не удалось получить значения по курсам!')
    return currencies


def get_currency(currency_url):
    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,"
                  "*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "referer": "https://yandex.ru/",
        "accept-encoding": "gzip, deflate, br",
        "user-agent": USER_AGENT
    }
    response = get_response(url=currency_url, headers=headers)
    if not response:
        return None

    currencies = pars_currency(response)
    return currencies


def make_excel_headers(ws):
    for col, val in enumerate(['Дата', 'Доллар', 'Изменение', 'Дата', 'Евро', 'Изменение', 'Евро/Доллар'], start=1):
        ws.cell(row=1, column=col).value = val

    for col_num, column in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G'], start=1):
        if col_num == 7:
            ws.column_dimensions[column].width = 20
        else:
            ws.column_dimensions[column].width = 15
        ws[f'{column}1'].font = Font(bold=True)
        ws[f'{column}1'].alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')

def send_mail(self):
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.ehlo()

    server.login('yklshn@gmail.com', 'password')
    subject = 'MOEX Parser'
    body = 'Файл содержит 10 строк'
    server.sendmail('admin@admin.ru', 'yklshn@gmail.com', body)
    server.quit()


def save_to_xlsx(fn, usd_currencies, eur_currencies):
    try:
        wb = Workbook()
        ws = wb.active

        make_excel_headers(ws)

        col_offset = 0

        for currency in [usd_currencies, eur_currencies]:
            last_row = 1
            for row in currency:
                last_row += 1

                try:
                    ratio = eur_currencies[last_row - 2].get("value") / usd_currencies[last_row - 2].get("value")
                    ws.cell(last_row, 7).value = round(ratio, RATIO_ROUND_DIGIT)
                    ws.cell(last_row, 7).alignment = Alignment(horizontal='center')
                    #ws.cell(last_row, 7).number_format = BUILTIN_FORMATS[FORMAT_NUMBER]
                except Exception as e:
                    print(f'Ошибка получения соотношения для строки №{last_row}: {e}')

                ws.cell(last_row, 1 + col_offset).value = row.get("date")
                ws.cell(last_row, 1 + col_offset).alignment = Alignment(horizontal='center')

                ws.cell(last_row, 2 + col_offset).value = row.get("value")
                ws.cell(last_row, 2 + col_offset).alignment = Alignment(horizontal='center')

                ws.cell(last_row, 3 + col_offset).value = row.get("change")
                ws.cell(last_row, 3 + col_offset).alignment = Alignment(horizontal='center')

            col_offset += 3

        wb.save(fn)
        print(f'Файл сохранен в "{fn}"')
        os.system('start excel.exe "{}"'.format(fn))
        fn.send_mail()
        return True
    except Exception as e:
        print(f'Ошибка сохранения: "{e}"!')
        return False


def main():
    print('Получение информации для доллара...')
    usd_currencies = get_currency(USD_URL)
    if not usd_currencies:
        return

    time.sleep(PAUSE_NEXT)
    print('Получение информации для евро...')
    eur_currencies = get_currency(EUR_URL)
    if not eur_currencies:
        return

    save_to_xlsx(FN_RESULT.format(datetime.datetime.now().strftime("%Y%m%d_%H%M%S")),
                 usd_currencies,
                 eur_currencies
                 )


if __name__ == '__main__':
    start_time = time.time()
    main()
    print(f"Работа завершена! Затраченное время {round(time.time() - start_time)} сек.")
